#!/usr/bin/env python3
"""Fallback for Workshop 6 if the agent hangs. Synthetic only.

The sheet is wrong on purpose: it omits the 400 mg cap so the room
still has something to catch in formula view.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT = Path(__file__).parent / "lab2_out"
OUT.mkdir(exist_ok=True)


def dosing():
    wb = Workbook()
    ws = wb.active
    ws.title = "teachicillin"
    headers = ["weight_kg", "age_months", "teachicillin_mg_per_dose", "doses_per_day", "notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    rows = [(1.5, 0), (3.2, 1), (8.0, 6), (14.0, 24), (45.0, 144)]
    for i, (wt, age) in enumerate(rows, start=2):
        ws.cell(i, 1, wt)
        ws.cell(i, 2, age)
        # Intentional miss: no TOO_SMALL branch and no 400 mg cap.
        ws.cell(i, 3, f"=A{i}*10")
        ws.cell(i, 4, 3)
        ws.cell(i, 5, "synthetic row")
    ws["A10"] = "NOT FOR CLINICAL USE — fictional drug — missing max cap on purpose"
    wb.save(OUT / "dosing.xlsx")
    (OUT / "README.txt").write_text(
        "Not for clinical use. Fictional teachicillin. Formula view: dose is weight*10 with no 400 mg cap.\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    dosing()
    print("wrote", list(OUT.iterdir()))
